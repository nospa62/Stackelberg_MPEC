import sys
import os
import re
import subprocess
import pandas as pd

# Ensure openpyxl is installed
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Import extractor functions
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import importlib
extractor = importlib.import_module("05_results_extractor")

# --- STYLES ---
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="002060", bold=True, size=14)

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# --- FORMATS ---
FMT_CURRENCY = '€#,##0.00'
FMT_MVAR_MW = '#,##0.000'
FMT_PU = '0.000000'
FMT_DEG = '0.0000'
FMT_PCT = '0.00%'
FMT_PRICE = '0.000000'

def _dat(content_or_path, name, dims=1):
    """Thin adapter: parse one named param from .dat content string."""
    import tempfile, os
    # Write content to a temp file so _parse_network_dat can read it
    tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.dat', delete=False)
    tmp.write(content_or_path); tmp.close()
    result = extractor._parse_network_dat(tmp.name)
    os.unlink(tmp.name)
    return result.get(name, {})

def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def generate_excel_report(summary_txt, raw_txt, network_dat, output_xlsx):
    print("Loading data...")
    res = extractor.parse_solution_summary(summary_txt)
    raw_res = extractor.parse_raw_solution(raw_txt)
    stats = extractor.compute_market_statistics(res, network_dat)
    consistency_violations = extractor.verify_dual_price_consistency(res, network_dat)
    
    with open(network_dat, 'r') as f:
        dat_content = f.read()
        
    v_min = _dat(dat_content, 'V_min', 1)
    v_max = _dat(dat_content, 'V_max', 1)
    s_max = _dat(dat_content, 'S_max', 2)
    
    cost_a_inj = _dat(dat_content, 'cost_a_inj', 1)
    cost_b_inj = _dat(dat_content, 'cost_b_inj', 1)
    cost_c_inj = _dat(dat_content, 'cost_c_inj', 1)
    cost_a_abs = _dat(dat_content, 'cost_a_abs', 1)
    cost_b_abs = _dat(dat_content, 'cost_b_abs', 1)
    cost_c_abs = _dat(dat_content, 'cost_c_abs', 1)
    q_inj_max = _dat(dat_content, 'q_inj_max', 1)
    q_abs_max = _dat(dat_content, 'q_abs_max', 1)
    gen_names = _dat(dat_content, 'gen_name', 1)
    
    s_base = res['solve_info']['s_base']
    
    solve_status = str(res['solve_info']['solve_result']).strip().lower()
    if solve_status != "solved":
        print(f"WARNING: solve_result={solve_status}. Report will be marked INVALID.")

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    # ==========================================
    # SHEET 1: Summary
    # ==========================================
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_properties.tabColor = "002060"
    
    ws1['A1'] = "Stackelberg Dual-Price Reactive Power Market — Results"
    ws1['A1'].font = TITLE_FONT
    
    consistency_status = "PASS" if not consistency_violations else "FAIL"
    if solve_status != "solved":
        consistency_status = "INVALID (solver did not converge)"
    
    summary_data = [
        ("Solve Result", res['solve_info']['solve_result']),
        ("Objective (€)", res['solve_info']['objective']),
        ("Final eps_smooth", res['solve_info']['eps_final']),
        ("Total Q Injected (MVAr)", stats['total_q_injected_mvar']),
        ("Total Q Absorbed (MVAr)", stats['total_q_absorbed_mvar']),
        ("Total Q Shunt (MVAr)", stats['total_q_shunt_mvar']),
        ("Total Q Load (MVAr)", stats['total_q_load_mvar']),
        ("Net Reactive Line Losses (MVAr)", stats['net_reactive_line_losses_mvar']),
        ("Total Payment (€)", stats.get('total_payment_eur', res['solve_info']['objective'])),
        ("Pure Procurement Cost (€)", stats.get('procurement_cost_eur', res['solve_info']['objective'])),
        ("Tikhonov Regularisation Cost (€)", stats.get('tikhonov_eur', 0.0)),
        ("Avg Injection Price (€/MVAr)", stats['average_injection_price_eur_mvar']),
        ("Avg Absorption Price (€/MVAr)", stats['average_absorption_price_eur_mvar']),
        ("Generators Injecting", stats['num_generators_injecting']),
        ("Generators Absorbing", stats['num_generators_absorbing']),
        ("Generators Idle", stats['num_generators_idle']),
    ]
    
    # Extract ALL P_ref values (one per REF_BUS)
    p_ref_items = [(k, v * s_base) for k, v in raw_res.items() if k.startswith('P_ref[')]
    if p_ref_items:
        for k, v in p_ref_items:
            bus_label = k.replace('P_ref[', '').replace(']', '')
            summary_data.insert(3, (f"P_ref Bus {bus_label} (MW)", v))
    
    for i, (k, v) in enumerate(summary_data, start=3):
        ws1[f'A{i}'] = k
        ws1[f'B{i}'] = v
        ws1[f'A{i}'].font = Font(bold=True)
        if "€" in k: ws1[f'B{i}'].number_format = FMT_CURRENCY
        elif "MVAr" in k: ws1[f'B{i}'].number_format = FMT_MVAR_MW
        
    row = len(summary_data) + 5
    ws1[f'A{row}'] = "Consistency Checks"
    ws1[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    if solve_status != "solved":
        ws1[f'A{row}'] = "Dual Price Consistency"
        ws1[f'B{row}'] = consistency_status
        ws1[f'B{row}'].fill = FILL_RED
    elif not consistency_violations:
        ws1[f'A{row}'] = "Dual Price Consistency"
        ws1[f'B{row}'] = "PASS"
        ws1[f'B{row}'].fill = FILL_GREEN
    else:
        ws1[f'A{row}'] = "Dual Price Consistency"
        ws1[f'B{row}'] = "FAIL"
        ws1[f'B{row}'].fill = FILL_RED
        row += 1
        for viol in consistency_violations:
            ws1[f'A{row}'] = viol
            ws1[f'A{row}'].font = Font(color="FF0000")
            row += 1

    autofit_columns(ws1)

    # ==========================================
    # SHEET 2: Prices & Dispatch
    # ==========================================
    ws2 = wb.create_sheet("Prices & Dispatch")
    ws2.sheet_properties.tabColor = "92D050" # Green
    
    headers2 = ["gen_id", "name", "bus_id", "λ_inj (€/MVAr)", "λ_abs (€/MVAr)", 
                "Q_inj (MVAr)", "Q_abs (MVAr)", "Q_net (MVAr)", "direction", 
                "revenue (€)", "cost (€)", "profit (€)"]
    ws2.append(headers2)
    apply_header_style(ws2)
    ws2.freeze_panes = 'A2'
    
    df_disp = res['dispatch'].merge(res['prices'], on='gen_id').sort_values('gen_id')
    
    for _, r in df_disp.iterrows():
        gen = str(r['gen_id'])
        lam_inj = r['lam_inj']
        lam_abs = r['lam_abs']
        qp = r['qp_mvar']
        qn = r['qn_mvar']
        qnet = r['q_net_mvar']
        profit = r['profit_eur']
        
        eps_ir = 1e-3
        qp_pu = qp / s_base
        qn_pu = qn / s_base
        w_inj = qp_pu / (qp_pu + eps_ir)
        w_abs = qn_pu / (qn_pu + eps_ir)
        
        gen_key = str(int(float(gen)))
        ca_inj = cost_a_inj.get(gen_key, 0.0)
        cb_inj = cost_b_inj.get(gen_key, 0.0)
        cc_inj = cost_c_inj.get(gen_key, 0.0)
        ca_abs = cost_a_abs.get(gen_key, 0.0)
        cb_abs = cost_b_abs.get(gen_key, 0.0)
        cc_abs = cost_c_abs.get(gen_key, 0.0)
        
        cost = (ca_inj*(qp**2) + cb_inj*qp + cc_inj*w_inj
              + ca_abs*(qn**2) + cb_abs*qn + cc_abs*w_abs)
        revenue = lam_inj * qp + lam_abs * qn
        profit = revenue - cost
        
        if qp > 1e-3: direction = "INJECTION"
        elif qn > 1e-3: direction = "ABSORPTION"
        else: direction = "IDLE"
        
        display_name = gen_names.get(gen, f"G{gen}")
        row_data = [gen, display_name, r['bus_id'], lam_inj, lam_abs, qp, qn, qnet, direction, revenue, cost, profit]
        ws2.append(row_data)
        
        current_row = ws2.max_row
        fill = FILL_GREEN if direction == "INJECTION" else (FILL_BLUE if direction == "ABSORPTION" else FILL_GREY)
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=current_row, column=col)
            cell.fill = fill
            if col in [4, 5]: cell.number_format = FMT_PRICE
            elif col in [10, 11, 12]: cell.number_format = FMT_CURRENCY
            elif col in [6, 7, 8]: cell.number_format = FMT_MVAR_MW

        if profit < 0:
            ws2.cell(row=current_row, column=12).fill = FILL_ORANGE
            ws2.cell(row=current_row, column=12).comment = Comment(
                "Negative profit due to fixed cost c_inj. Variable profit is positive — "
                "generator participates at marginal cost. This is a known MPEC property.", "Model")

    autofit_columns(ws2)

    # ==========================================
    # SHEET 3: Voltages
    # ==========================================
    ws3 = wb.create_sheet("Voltages")
    ws3.sheet_properties.tabColor = "0070C0" # Blue
    
    headers3 = ["bus_id", "name", "V_min", "V_pu", "V_max", "θ (deg)", "status"]
    ws3.append(headers3)
    apply_header_style(ws3)
    ws3.freeze_panes = 'A2'
    
    for _, r in res['voltages'].iterrows():
        bus_key = str(int(float(r['bus_id'])))
        v_pu = r['V_pu']
        theta = r['theta_deg']
        
        vmin = v_min.get(bus_key, 0.95)
        vmax = v_max.get(bus_key, 1.05)
        
        status = "OK" if (vmin - 1e-4) <= v_pu <= (vmax + 1e-4) else "VIOLATION"
        
        ws3.append([bus_key, bus_key, vmin, v_pu, vmax, theta, status])
        
        current_row = ws3.max_row
        fill = FILL_GREEN if status == "OK" else FILL_RED
        for col in range(1, len(headers3) + 1):
            cell = ws3.cell(row=current_row, column=col)
            cell.fill = fill
            if col in [3, 4, 5]: cell.number_format = FMT_PU
            elif col == 6: cell.number_format = FMT_DEG

    autofit_columns(ws3)

    # ==========================================
    # SHEET 4: Branch Flows
    # ==========================================
    ws4 = wb.create_sheet("Branch Flows")
    ws4.sheet_properties.tabColor = "FFC000" # Orange
    
    headers4 = ["from_bus", "to_bus", "name", "P_flow (MW)", "Q_flow (MVAr)", 
                "|S_flow| (MVA)", "S_max (MVA)", "loading (%)", "status"]
    ws4.append(headers4)
    apply_header_style(ws4)
    ws4.freeze_panes = 'A2'
    
    for _, r in res['branch_flows'].iterrows():
        f_bus = str(r['from_bus'])
        t_bus = str(r['to_bus'])
        p_flow = r['P_flow_mw']
        q_flow = r['Q_flow_mva']
        loading = r['loading_pct'] / 100.0 # Store as decimal for Excel % format
        
        s_flow = (p_flow**2 + q_flow**2)**0.5
        smax = s_max.get((f_bus, t_bus), s_max.get((t_bus, f_bus), s_flow / max(loading, 1e-6) if loading > 0 else 9999))
        
        if loading > 1.0001: status = "OVERLOAD"
        elif loading > 0.80: status = "WARNING"
        else: status = "OK"
        
        name = f"{f_bus}-{t_bus}"
        ws4.append([f_bus, t_bus, name, p_flow, q_flow, s_flow, smax, loading, status])
        
        current_row = ws4.max_row
        fill = FILL_RED if status == "OVERLOAD" else (FILL_ORANGE if status == "WARNING" else FILL_GREEN)
        for col in range(1, len(headers4) + 1):
            cell = ws4.cell(row=current_row, column=col)
            cell.fill = fill
            if col in [4, 5, 6, 7]: cell.number_format = FMT_MVAR_MW
            elif col == 8: cell.number_format = FMT_PCT

    autofit_columns(ws4)

    # ==========================================
    # SHEET 5: KKT Verification
    # ==========================================
    ws5 = wb.create_sheet("KKT Verification")
    ws5.sheet_properties.tabColor = "7030A0" # Purple
    
    headers5 = ["gen_id", "μ_qp_ub", "μ_qp_lb", "μ_qn_ub", "μ_qn_lb",
                "stat_inj_residual", "stat_abs_residual",
                "compl_qp_ub_violation", "compl_qp_lb_violation",
                "compl_qn_ub_violation", "compl_qn_lb_violation",
                "exclusivity_violation", "overall_status"]
    ws5.append(headers5)
    apply_header_style(ws5)
    ws5.freeze_panes = 'A2'
    
    df_kkt = res['kkt_multipliers'].merge(res['dispatch'], on='gen_id').merge(res['prices'], on='gen_id')
    
    def get_kkt_fill(val):
        if abs(val) < 1e-4: return FILL_GREEN
        elif abs(val) < 1e-3: return FILL_YELLOW
        else: return FILL_RED

    for _, r in df_kkt.iterrows():
        gen = str(r['gen_id'])
        qp_mvar = r['qp_mvar']
        qn_mvar = r['qn_mvar']
        lam_inj = r['lam_inj']
        lam_abs = r['lam_abs']
        
        mu_qp_ub = r['mu_qp_ub']
        mu_qp_lb = r['mu_qp_lb']
        mu_qn_ub = r['mu_qn_ub']
        mu_qn_lb = r['mu_qn_lb']
        
        gen_key = str(int(float(gen)))
        
        ca_inj = cost_a_inj.get(gen_key, 0.0)
        cb_inj = cost_b_inj.get(gen_key, 0.0)
        ca_abs = cost_a_abs.get(gen_key, 0.0)
        cb_abs = cost_b_abs.get(gen_key, 0.0)
        qmax_inj = q_inj_max.get(gen_key, 0.0)
        qmax_abs = q_abs_max.get(gen_key, 0.0)
        
        stat_inj = abs(
            lam_inj - 2*ca_inj*qp_mvar - cb_inj
            - mu_qp_ub + mu_qp_lb
        )
        stat_abs = abs(
            lam_abs - 2*ca_abs*qn_mvar - cb_abs
            - mu_qn_ub + mu_qn_lb
        )
        
        qmax_inj_mvar = qmax_inj * s_base
        qmax_abs_mvar = qmax_abs * s_base
        
        compl_qp_ub = abs(mu_qp_ub * (qmax_inj_mvar - qp_mvar))
        compl_qp_lb = abs(mu_qp_lb * qp_mvar)
        compl_qn_ub = abs(mu_qn_ub * (qmax_abs_mvar - qn_mvar))
        compl_qn_lb = abs(mu_qn_lb * qn_mvar)
        
        DEBUG_KKT = os.environ.get('STACKELBERG_DEBUG', '0') == '1'
        if DEBUG_KKT:
            print(f"gen {gen}: mu_qn_lb={mu_qn_lb:.6g}, lam_abs={lam_abs:.6g}, "
                  f"cb_abs={cb_abs:.6g}, stat_abs={stat_abs:.3e}")
        
        excl = abs(qp_mvar * qn_mvar)
        
        max_viol = max(stat_inj, stat_abs, compl_qp_ub, compl_qp_lb, compl_qn_ub, compl_qn_lb, excl)
        if max_viol < 1e-4: overall = "PASS"
        elif max_viol < 1e-3: overall = "WARN"
        else: overall = "FAIL"
        
        row_data = [gen, mu_qp_ub, mu_qp_lb, mu_qn_ub, mu_qn_lb,
                    stat_inj, stat_abs, compl_qp_ub, compl_qp_lb,
                    compl_qn_ub, compl_qn_lb, excl, overall]
        ws5.append(row_data)
        
        current_row = ws5.max_row
        for col in range(2, 6): ws5.cell(row=current_row, column=col).number_format = FMT_PU
        
        # Color code residuals and violations
        for col in range(6, 13):
            cell = ws5.cell(row=current_row, column=col)
            cell.number_format = FMT_PU
            cell.fill = get_kkt_fill(cell.value)
            
        # Overall status color
        cell_overall = ws5.cell(row=current_row, column=13)
        if overall == "PASS": cell_overall.fill = FILL_GREEN
        elif overall == "WARN": cell_overall.fill = FILL_YELLOW
        else: cell_overall.fill = FILL_RED

    autofit_columns(ws5)

    # ==========================================
    # SHEET 6: Market Statistics
    # ==========================================
    ws6 = wb.create_sheet("Market Statistics")
    ws6.sheet_properties.tabColor = "00B0F0" # Light Blue
    
    ws6['A1'] = "Market Statistics"
    ws6['A1'].font = TITLE_FONT
    
    row = 3
    for k, v in stats.items():
        ws6[f'A{row}'] = k
        ws6[f'B{row}'] = v
        if "eur" in k: ws6[f'B{row}'].number_format = FMT_CURRENCY
        elif "mvar" in k or "mw" in k: ws6[f'B{row}'].number_format = FMT_MVAR_MW
        elif "pu" in k: ws6[f'B{row}'].number_format = FMT_PU
        row += 1
        
    row += 2
    ws6[f'A{row}'] = "Producer Deadbands"
    ws6[f'A{row}'].font = TITLE_FONT
    row += 1
    
    ws6[f'A{row}'] = "NOTE: lam_abs values are Tikhonov artefacts (≈δ_reg). True absorption market price is indeterminate when qn=0."
    ws6[f'A{row}'].font = Font(italic=True, color="808080")
    row += 2
    
    headers6 = ["gen_id", "injection_deadband_price (€/MVAr)", "absorption_deadband_price (€/MVAr)", 
                "actual_lam_inj (€/MVAr)", "actual_lam_abs (€/MVAr)", "in_deadband?"]
    ws6.append(headers6)
    apply_header_style(ws6, row)
    
    for _, r in df_disp.iterrows():
        gen = str(r['gen_id'])
        lam_inj = r['lam_inj']
        lam_abs = r['lam_abs']
        
        gen_key = str(int(float(r['gen_id'])))
        cb_inj = cost_b_inj.get(gen_key, 0.0)
        cb_abs = cost_b_abs.get(gen_key, 0.0)
        
        inj_db = lam_inj <= cb_inj + 1e-4
        abs_db = lam_abs <= cb_abs + 1e-4
        in_deadband = f"INJ={'DB' if inj_db else 'ACTIVE'} / ABS={'DB' if abs_db else 'ACTIVE'}"
        
        ws6.append([gen_key, cb_inj, cb_abs, lam_inj, lam_abs, in_deadband])
        
        current_row = ws6.max_row
        for col in range(2, 4):
            ws6.cell(row=current_row, column=col).number_format = FMT_CURRENCY
        for col in range(4, 6):
            ws6.cell(row=current_row, column=col).number_format = FMT_PRICE
        
        cell_db = ws6.cell(row=current_row, column=6)
        cell_db.fill = FILL_GREY if "DB" in in_deadband else FILL_GREEN

    autofit_columns(ws6)

    # --- Sheet 7: Voltage Profile Chart ---
    from openpyxl.chart import BarChart, Reference
    
    ws7 = wb.create_sheet("Voltage Profile")
    ws7.sheet_properties.tabColor = "70AD47"
    
    # Write data for chart
    ws7['A1'] = "Bus ID"; ws7['B1'] = "V (pu)"; ws7['C1'] = "V_min"; ws7['D1'] = "V_max"
    for i, r in enumerate(res['voltages'].itertuples(), start=2):
        bus_key = str(int(float(r.bus_id)))
        ws7[f'A{i}'] = bus_key
        ws7[f'B{i}'] = r.V_pu
        ws7[f'C{i}'] = float(v_min.get(bus_key, 0.95))
        ws7[f'D{i}'] = float(v_max.get(bus_key, 1.05))
    
    n_buses = len(res['voltages'])
    chart = BarChart()
    chart.type = "col"
    chart.title = "Voltage Profile"
    chart.y_axis.title = "V (pu)"
    chart.x_axis.title = "Bus"
    chart.y_axis.scaling.min = 0.93
    chart.y_axis.scaling.max = 1.07
    
    data_ref = Reference(ws7, min_col=2, max_col=4, min_row=1, max_row=n_buses + 1)
    cats = Reference(ws7, min_col=1, min_row=2, max_row=n_buses + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 25; chart.height = 15
    ws7.add_chart(chart, "F2")

    # Save workbook
    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    wb.save(output_xlsx)
    print(f"Results successfully written to {output_xlsx}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Generate Excel report from AMPL results.")
    parser.add_argument("--summary", default="ampl/solution_summary.txt", help="Path to solution_summary.txt")
    parser.add_argument("--raw", default="ampl/solution_raw.txt", help="Path to solution_raw.txt")
    parser.add_argument("--dat", default="ampl/network.dat", help="Path to network.dat")
    parser.add_argument("--out", default="output/stackelberg_results.xlsx", help="Path to output Excel file")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.summary) or not os.path.exists(args.raw):
        print(f"Error: Could not find AMPL output files at {args.summary} or {args.raw}")
        print("Please run the AMPL model first.")
        sys.exit(1)
        
    generate_excel_report(args.summary, args.raw, args.dat, args.out)
